# Pyhton 下用Django语法增删查改，注意创建空的数据库需要在终端或数据库软件中，表结构需要在models.py模块
import os
import django
os.environ.setdefault(
    "DJANGO_SETTINGS_MODULE",
    "mysite.settings")  # project_name 项目名称
django.setup()
from django.core.exceptions import ObjectDoesNotExist
from dateutil.relativedelta import relativedelta
from tkinter import filedialog
import tkinter as tk
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
from openpyxl import load_workbook, Workbook
import pymysql
from collector import models, models_view
import datetime
import time
import inspect
import sys


# 后续开发：用户登录，封装成一个函数，调用传入用户名和密码即可：


# 从EXCEL中导入数据,封装成一个函数，调用是传入表格路径和需要导入的表名即可
def xlstodb(localpath, sheetlist):
    wb = load_workbook(localpath, data_only=True)  # 不导入公式，只导入值
    ws = wb['Cover']
    mon = ws['C5'].value
    ver = ws['B7'].value
    Last_date = datetime.date.today() - relativedelta(months=1)
    if mon == None or ver == None:
        print("请用此文件先执行'填报'程序获取可上传版本号后再执行'上传'程序！")
    else:
        if int(mon) == datetime.datetime.now().month or int(mon) == Last_date.month:  # 只能上传本月或上月报表:
            try:
                models.Version.objects.get(versionname=str(ver)) #对照数据库检查Excel里的版本号是否存在
                for tbname, obj in inspect.getmembers(models):
                    if inspect.isclass(obj) and tbname in sheetlist:
                        fieldlist = []  # 临时存放某一张表的所有字段名
                        for fieldobj in obj._meta.fields:
                            fieldlist.append(fieldobj.__dict__['attname']) #数据库字段名而非models字段名
                        ws = wb[tbname]
                        rowdict = {}  # 存储单行数据
                        rowlist = []  # 存储多行数据
                        rows = ws.max_row
                        columns = ws.max_column
                        for row in range(3, rows + 1):
                            for column in range(1, columns + 1):
                                rowdict[fieldlist[column]] = str(
                                    ws.cell(row=row, column=column).value) #适用于数据库字段比excel字段多一个ID：auto_created
                                # print('程序正在执行单元格识别……')
                            objing = obj(**rowdict)
                            rowlist.append(objing)
                            rowdict = {}
                        print(tbname, '程序正在执行行数据导入……')
                        try:
                            obj.objects.bulk_create(rowlist)  # 使用bulk_create批量导入
                            print(tbname + '批量导入成功！')
                        except Exception as e:
                            print(tbname + '导入失败：', e)
                    wb.close()  # 关闭excel
            except models.Version.objects.DoesNotExist:
                print("该文件版本号不存在，请用此文件执行'填报'程序更新后再执行'上传'程序！")
        else:
            print(
                "此工作簿'会计期间为%s月，系统要求只能上传本月或上月报表！请退出后直接用本工作簿执行'填报'程序后再'上传'！" %
                (mon))




application_window = tk.Tk()
application_window.withdraw()  # 隐藏消息框
my_filetypes = [('xlsx files', '.xlsx')] # Build a list of tuples for each file type the file dialog should display
localpath = filedialog.askopenfilename(
    parent=application_window,
    initialdir=os.getcwd(),
    title="请选择上传数据文件:",
    filetypes=my_filetypes)
sheetlist = ['ActualData', 'BudgetData', 'YtdData']
xlstodb(localpath, sheetlist)
