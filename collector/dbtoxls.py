# Pyhton 下用Django语法增删查改，注意创建空的数据库需要在终端或数据库软件中，表结构需要在models.py模块
import os
import django
import datetime
import time
import inspect
import sys
os.environ.setdefault(
    "DJANGO_SETTINGS_MODULE",
    "mysite.settings")  # project_name 项目名称
django.setup()
from tkinter import filedialog
import tkinter as tk
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
from openpyxl import load_workbook, Workbook
import pymysql
from collector import models, models_view


application_window = tk.Tk()
application_window.withdraw()  # 隐藏消息框

modellist = [models,models_view]
modelnum = len(modellist)
# 一段循环写完所有表
wb = Workbook()
i = 0
while i < modelnum:
    for name, obj in inspect.getmembers(modellist[i]):
        if inspect.isclass(obj):
            ws = wb.create_sheet('%s' % (name))
            # 导出表结构
            fielddic = {}
            for field in obj._meta.fields:
                if field.__dict__['auto_created'] == False:  # 系统自动增长的字段ID字段不导出
                    fielddic[field.attname] = field.verbose_name
                # print(fielddic)
            fields1 = list(fielddic)
            fields2 = list(fielddic.values())
            # print(fields1)
            ws.append(fields1)
            ws.append(fields2)
            # 导出表内容
            for rows in obj.objects.all(
            ).values_list():  # 未排除自动增长的字段内容导出，结构没有，内容有（小BUG）
                ws.append(list(rows))
            ws.row_dimensions.group(1, 1, hidden=True)
    i = i+1

ws = wb.active
ws.title = 'Cover'
ws['A4'] = '填报单位：'
ws['A4'].font = Font(name='Arial', size=14, color=colors.RED, italic=True)
ws['A5'] = '填报期间:'
ws['A6'] = '币种：'
ws['A7'] = '版本号'


application_window = tk.Tk()
application_window.withdraw()  # 隐藏消息框

# Build a list of tuples for each file type the file dialog should display
my_filetypes = [('xlsx files', '.xlsx')]
# localpath = filedialog.asksaveasfile()
# localpath = filedialog.askdirectory()
localpath = filedialog.asksaveasfilename(
    parent=application_window,
    initialdir=os.getcwd(),
    title="Please select a location:",
    filetypes=my_filetypes)


wb.save(localpath)
